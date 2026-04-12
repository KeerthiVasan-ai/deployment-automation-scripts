import firebase_admin
from firebase_admin import credentials, firestore, auth
from openpyxl import load_workbook
from datetime import datetime

# Initialize Firebase
cred = credentials.Certificate("D:\\deployment-automation-scripts\\SmartClassAutomation\\smartreserve.json")
firebase_admin.initialize_app(cred)
db = firestore.client()

# Load Excel file
workbook = load_workbook('D:\\deployment-automation-scripts\\SmartClassAutomation\\CSE_STAFF_DETAILS.xlsx')
worksheet = workbook.active

# Get the previous year
current_year = datetime.now().year
last_year = current_year - 1

# Log file name
log_file = "old_bookings_log_06_07.txt"

def log_and_delete_old_bookings():
    with open(log_file, "a") as log:
        log.write(f"\n===== Log for {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} =====\n")

        for row in worksheet.iter_rows(values_only=True):  
            email = row[2]  
            if email:
                try: 
                    user = auth.get_user_by_email(email)
                    uid = user.uid
                    print(f"Checking bookings for: {email} (UID: {uid})")

                    bookings = db.collection("bookingUserDetails").document(uid).collection("bookings").get()

                    for doc in bookings:
                        booking_data = doc.to_dict()
                        if "date" in booking_data and isinstance(booking_data["date"], str):
                            try:
                                # booking_date = datetime.strptime(booking_data["date"], "%d-%m-%Y")
                                
                                # if booking_date.year == last_year and 1 <= booking_date.month <= 6:
                                log.write(f"Old Booking Found: {doc.id} -> {booking_data}\n")
                                print(f"Logging and deleting old booking: {doc.id} -> {booking_data}")

                                db.collection("bookingUserDetails").document(uid).collection("bookings").document(doc.id).delete()
                                print(f"Deleted old booking: {doc.id}")

                            except ValueError:
                                log.write(f"Invalid date format in document {doc.id}: {booking_data['date']}\n")
                                print(f"Invalid date format in document {doc.id}: {booking_data['date']}")

                except auth.UserNotFoundError:
                    log.write(f"User with email {email} not found in Firebase Authentication.\n")
                    print(f"User with email {email} not found in Firebase Authentication.")

log_and_delete_old_bookings()
