import firebase_admin
from firebase_admin import credentials, firestore, auth

cred = credentials.Certificate("smartreserve.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

from openpyxl import load_workbook

workbook = load_workbook('CSE_STAFF_DETAILS.xlsx')
worksheet = workbook.active

def create_allotted_slots():
    for row in worksheet.iter_rows(values_only=True):
        email = row[2]
        if email != None:
            user = auth.get_user_by_email(email)
            uid = user.uid
            print(uid)
            doc_ref = db.collection("allottedSlots").document(uid)
            doc_ref.set({
                "allottedSlots": 2
            })

    print("Document created successfully!")

create_allotted_slots()
