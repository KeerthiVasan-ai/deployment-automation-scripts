import firebase_admin
from firebase_admin import credentials,firestore,auth
from datetime import datetime,timedelta

cred = credentials.Certificate("smartreserve.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

from openpyxl import load_workbook

workbook = load_workbook('CSE_STAFF_DETAILS.xlsx')
worksheet = workbook.active

def update_individual_slot_key():
    for row in worksheet.iter_rows(values_only=True):
        email = row[2]
        if email != None:
            user = auth.get_user_by_email(email)
            uid = user.uid
            print(uid)
            documents = db.collection("bookingUserDetails").document(uid).collection("bookings").get()
            for document in documents:
                data = document.to_dict()
                time = data['slots'][0]
                key = key_to_time(time)
                print(time,":",key)
                booking_ref = db.collection("bookingUserDetails").document(uid).collection("bookings").document(document.id)
                booking_ref.update({
                    "slotKey":key
                })

def update_date_slot_key():
    start_date = datetime(2024, 2, 1)
    end_date = datetime(2024, 4, 30)
    date_list = []

    current_date = start_date

    while current_date <= end_date:
        date_list.append(current_date.strftime('%d-%m-%Y'))
        current_date += timedelta(days=1) 

    for date in date_list:
        documents = db.collection("bookingDetails").document(date).collection("booking").get()
        print(date,":",len(documents))
        for document in documents:
            data = document.to_dict()
            time = data['slots'][0]
            key = key_to_time(time)
            print(time,":",key)
            booking_ref = db.collection("bookingDetails").document(date).collection("booking").document(document.id)
            booking_ref.update({
                "slotKey":key
            })
    

def key_to_time(time):
    mapper = {
        "08:30AM": '1',
        "09:15AM": '2',
        "10:00AM": '3',
        "10:55AM": '4',
        "11:40AM": '5',
        "01:30PM": '6',
        "02:15PM": '7',
        "03:00PM": '8',
        "03:45PM": '9'
    }
    return mapper[time]

if __name__ == '__main__':
    update_individual_slot_key()
    update_date_slot_key()