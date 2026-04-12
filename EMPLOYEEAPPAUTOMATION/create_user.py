import firebase_admin
from firebase_admin import auth

cred = firebase_admin.credentials.Certificate('D:\SmartClassAutomation\EMPLOYEEAPPAUTOMATION\cse-emp-app.json')
firebase_admin.initialize_app(cred)

from openpyxl import load_workbook

workbook = load_workbook('D:\SmartClassAutomation\CSE_STAFF_DETAILS.xlsx')
worksheet = workbook.active

for row in worksheet.iter_rows(values_only=True):
    email = row[2]
    if email != None:
        user = auth.create_user(
            email=email,
            password='12345678',
        )
        print(user.uid)