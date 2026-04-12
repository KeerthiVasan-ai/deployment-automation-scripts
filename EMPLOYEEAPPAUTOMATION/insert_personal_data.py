import firebase_admin
from firebase_admin import auth,firestore
import pandas as pd
from openpyxl import load_workbook

cred = firebase_admin.credentials.Certificate('D:\\Automation Scripts\\EMPLOYEEAPPAUTOMATION\\cse-emp-app.json')
firebase_admin.initialize_app(cred)

db = firestore.client()

csv_df = pd.read_csv("AU_employee_data.csv")

workbook = load_workbook('D:\\Automation Scripts\\SmartClassAutomation\\CSE_STAFF_DETAILS.xlsx')
worksheet = workbook.active

def retrieve_personal_data(token_number):
    data = csv_df[csv_df["tokenNumber"] == token_number]
    return data[
        ["tokenNumber", "name","designation","qualification","dob","specialization",
         "presentAddress","phoneNumber"]
    ]

def retrieve_academic_data(token_number):
    data = csv_df[csv_df["tokenNumber"] == token_number]
    return data[
        ["joiningDate","meAwarded","meGuide","phdAwarded","phdGuided"
        #   "attended_national","attended_international","attended_seminar","attended_symposia","attended_workshop",
        #   "conducted_national","conducted_international","conducted_seminar","conducted_symposia","conducted_workshop",
        #   "completed_major_project","completed_minor_project"
        ]
    ]



for row in worksheet.iter_rows(values_only = True):
    email = row[2]
    if email != None:
        user = auth.get_user_by_email(email)
        uid = user.uid
        print(uid)
        data = retrieve_academic_data(row[1])
        print(data.to_dict(orient="records")[0])
        db.collection("academic_details").document(uid).set(data.to_dict(orient='records')[0])
